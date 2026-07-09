from datetime import date, timedelta

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from .models import (
    RoleMatrix, Skill, SkillBenchmark, SkillMatrix, EmployeeSkill, EmployeeSkillHistory,
    DevelopmentPlan, user_can_manage_employee,
)
from .views import safe_json
from .reports import build_team_report_data


class GapScoringTests(TestCase):
    def setUp(self):
        self.role = RoleMatrix.objects.create(title='Engineer')
        self.mandatory_skill = Skill.objects.create(name='Python')
        self.mandatory_skill_2 = Skill.objects.create(name='SQL')
        self.optional_skill = Skill.objects.create(name='Public Speaking')

        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.mandatory_skill, required_level=4, is_mandatory=True)
        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.mandatory_skill_2, required_level=3, is_mandatory=True)
        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.optional_skill, required_level=3, is_mandatory=False)

        self.employee = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=self.role)
        EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.mandatory_skill, actual_level=1)  # gap 3, mandatory
        EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.mandatory_skill_2, actual_level=3)  # gap 0, mandatory
        EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.optional_skill, actual_level=0)  # gap 3, optional

    def test_get_skill_gap_data_computes_gap_and_weight(self):
        gaps = {g['skill'].name: g for g in self.employee.get_skill_gap_data()}
        self.assertEqual(gaps['Python']['gap'], 3)
        self.assertEqual(gaps['Python']['weight'], 2)
        self.assertEqual(gaps['Public Speaking']['gap'], 3)
        self.assertEqual(gaps['Public Speaking']['weight'], 1)

    def test_overall_gap_score_weights_mandatory_skills_higher(self):
        # weighted gaps: Python 3*2=6, SQL 0*2=0, Public Speaking 3*1=3; total weight 2+2+1=5
        self.assertAlmostEqual(self.employee.get_overall_gap_score(), 9 / 5)

    def test_skills_met_percentage_is_unweighted(self):
        # 1 of 3 skills met (SQL)
        self.assertEqual(self.employee.get_skills_met_percentage(), round(1 / 3 * 100))

    def test_no_benchmarks_returns_zero(self):
        lone_employee = SkillMatrix.objects.create(name='No Role')
        self.assertEqual(lone_employee.get_overall_gap_score(), 0)
        self.assertEqual(lone_employee.get_skills_met_percentage(), 0)


class SkillHistoryTests(TestCase):
    def setUp(self):
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace')
        self.python = Skill.objects.create(name='Python')
        self.sql = Skill.objects.create(name='SQL')

    def _change_level(self, skill, new_level):
        """Create-then-update so EmployeeSkill.save() creates a history row (skips on create)."""
        es, _ = EmployeeSkill.objects.get_or_create(skill_matrix=self.employee, skill=skill, defaults={'actual_level': 0})
        es.actual_level = new_level
        es.save()
        return es

    def test_skill_with_multiple_changes_returns_ordered_points(self):
        self._change_level(self.python, 3)  # history row: level 3
        es = self._change_level(self.python, 4)  # history row: level 4

        rows = list(EmployeeSkillHistory.objects.filter(skill_matrix=self.employee, skill=self.python).order_by('pk'))
        EmployeeSkillHistory.objects.filter(pk=rows[0].pk).update(evaluated_on='2026-01-10T00:00:00Z')
        EmployeeSkillHistory.objects.filter(pk=rows[1].pk).update(evaluated_on='2026-01-20T00:00:00Z')
        # Current row's last_evaluated coincides with the latest history point, so it collapses
        # into it rather than adding a synthetic third point.
        EmployeeSkill.objects.filter(pk=es.pk).update(last_evaluated=date(2026, 1, 20))

        history = self.employee.get_skill_history()
        self.assertEqual(history['Python'], [
            {'date': '2026-01-10', 'level': 3},
            {'date': '2026-01-20', 'level': 4},
        ])

    def test_skill_with_no_history_rows_still_returns_current_level(self):
        EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.sql, actual_level=2)
        EmployeeSkill.objects.filter(skill_matrix=self.employee, skill=self.sql).update(last_evaluated=date(2026, 3, 1))

        history = self.employee.get_skill_history()
        self.assertEqual(history['SQL'], [{'date': '2026-03-01', 'level': 2}])

    def test_same_day_history_points_are_deduped_keeping_the_latest(self):
        EmployeeSkillHistory.objects.create(skill_matrix=self.employee, skill=self.python, recorded_level=3)
        EmployeeSkillHistory.objects.create(skill_matrix=self.employee, skill=self.python, recorded_level=4)
        rows = list(EmployeeSkillHistory.objects.filter(skill_matrix=self.employee, skill=self.python).order_by('pk'))
        for row in rows:
            EmployeeSkillHistory.objects.filter(pk=row.pk).update(evaluated_on='2026-01-10T00:00:00Z')
        # No EmployeeSkill row at all for this skill, so only the two history rows are in play.

        history = self.employee.get_skill_history()
        self.assertEqual(history['Python'], [{'date': '2026-01-10', 'level': 4}])


class ViewAccessControlTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin', password='pass12345', is_staff=True)
        self.plain_user = User.objects.create_user('manager', password='pass12345')
        self.role = RoleMatrix.objects.create(title='Engineer')
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=self.role)

    def test_anonymous_is_redirected_to_login(self):
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

        response = self.client.get(reverse('employee_list'))
        self.assertEqual(response.status_code, 302)

    def test_logged_in_user_can_view_dashboard(self):
        self.client.force_login(self.plain_user)
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)

    def test_non_staff_forbidden_from_delete_view(self):
        self.client.force_login(self.plain_user)
        response = self.client.get(reverse('employee_delete', args=[self.employee.id]))
        self.assertEqual(response.status_code, 403)

    def test_staff_can_access_delete_view(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse('employee_delete', args=[self.employee.id]))
        self.assertEqual(response.status_code, 200)

    def test_non_staff_forbidden_from_skill_update_endpoint(self):
        self.client.force_login(self.plain_user)
        response = self.client.post(reverse('employee_skill_update', args=[self.employee.id]), {})
        self.assertEqual(response.status_code, 403)


class DevelopmentPlanModelTests(TestCase):
    def setUp(self):
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace')
        self.skill = Skill.objects.create(name='Python')

    def _plan(self, **overrides):
        defaults = {'skill_matrix': self.employee, 'skill': self.skill, 'action': 'Take a course'}
        defaults.update(overrides)
        return DevelopmentPlan.objects.create(**defaults)

    def test_is_overdue_true_for_past_target_date_still_open(self):
        plan = self._plan(target_date=timezone.now().date() - timedelta(days=1), status='in_progress')
        self.assertTrue(plan.is_overdue)

    def test_is_overdue_false_for_future_target_date(self):
        plan = self._plan(target_date=timezone.now().date() + timedelta(days=1), status='in_progress')
        self.assertFalse(plan.is_overdue)

    def test_is_overdue_false_when_completed_even_if_past_due(self):
        plan = self._plan(target_date=timezone.now().date() - timedelta(days=1), status='completed')
        self.assertFalse(plan.is_overdue)

    def test_is_overdue_false_without_target_date(self):
        plan = self._plan(target_date=None, status='in_progress')
        self.assertFalse(plan.is_overdue)


class DevelopmentPlanAccessControlTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin', password='pass12345', is_staff=True)
        self.plain_user = User.objects.create_user('manager', password='pass12345')
        self.role = RoleMatrix.objects.create(title='Engineer')
        self.skill = Skill.objects.create(name='Python')
        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.skill, required_level=4)
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=self.role)
        self.plan = DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course',
        )

    def test_plain_user_can_view_plan_list_and_profile_plans_tab(self):
        self.client.force_login(self.plain_user)
        self.assertEqual(self.client.get(reverse('development_plan_list')).status_code, 200)
        self.assertEqual(self.client.get(reverse('employee_profile', args=[self.employee.id])).status_code, 200)

    def test_non_staff_forbidden_from_creating_a_plan(self):
        self.client.force_login(self.plain_user)
        response = self.client.get(reverse('development_plan_add', args=[self.employee.id]))
        self.assertEqual(response.status_code, 403)

    def test_staff_can_create_update_and_delete_a_plan(self):
        self.client.force_login(self.staff_user)

        response = self.client.post(reverse('development_plan_add', args=[self.employee.id]), {
            'skill': self.skill.id,
            'action': 'Pair with a senior dev',
            'status': 'not_started',
        })
        self.assertEqual(response.status_code, 302)
        new_plan = DevelopmentPlan.objects.get(action='Pair with a senior dev')
        self.assertEqual(new_plan.created_by, self.staff_user)

        response = self.client.post(reverse('development_plan_update', args=[new_plan.id]), {
            'skill': self.skill.id,
            'action': 'Pair with a senior dev',
            'status': 'in_progress',
        })
        self.assertEqual(response.status_code, 302)
        new_plan.refresh_from_db()
        self.assertEqual(new_plan.status, 'in_progress')

        response = self.client.post(reverse('development_plan_delete', args=[new_plan.id]))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(DevelopmentPlan.objects.filter(id=new_plan.id).exists())


class SelfRatingModelTests(TestCase):
    def setUp(self):
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace')
        self.skill = Skill.objects.create(name='Python')

    def test_rating_gap_none_without_self_rating(self):
        es = EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.skill, actual_level=3)
        self.assertIsNone(es.rating_gap)

    def test_rating_gap_is_self_minus_actual(self):
        es = EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2, self_rated_level=4,
        )
        self.assertEqual(es.rating_gap, 2)

    def test_user_can_manage_employee_staff_always_true(self):
        staff = User.objects.create_user('staff1', password='pass12345', is_staff=True)
        self.assertTrue(user_can_manage_employee(staff, self.employee))

    def test_user_can_manage_employee_true_for_own_manager(self):
        manager_user = User.objects.create_user('mgr1', password='pass12345')
        manager_employee = SkillMatrix.objects.create(name='Manager Mike', user=manager_user)
        self.employee.manager = manager_employee
        self.employee.save()
        self.assertTrue(user_can_manage_employee(manager_user, self.employee))

    def test_user_can_manage_employee_false_for_unrelated_user(self):
        other_user = User.objects.create_user('other1', password='pass12345')
        SkillMatrix.objects.create(name='Someone Else', user=other_user)
        self.assertFalse(user_can_manage_employee(other_user, self.employee))


class SelfRatingWorkflowTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin2', password='pass12345', is_staff=True)
        self.employee_user = User.objects.create_user('employee1', password='pass12345')
        self.manager_user = User.objects.create_user('manager1', password='pass12345')
        self.other_user = User.objects.create_user('other_employee', password='pass12345')

        self.role = RoleMatrix.objects.create(title='Engineer')
        self.skill = Skill.objects.create(name='Python')
        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.skill, required_level=4)

        self.manager_employee = SkillMatrix.objects.create(name='Manager Mike', user=self.manager_user)
        self.employee = SkillMatrix.objects.create(
            name='Ada Lovelace', role_matrix=self.role, user=self.employee_user, manager=self.manager_employee,
        )
        self.unrelated_employee = SkillMatrix.objects.create(name='Someone Else', user=self.other_user)

    def test_employee_can_submit_own_self_rating(self):
        self.client.force_login(self.employee_user)
        response = self.client.post(
            reverse('employee_self_rating_update', args=[self.employee.id]),
            {'skill_id': self.skill.id, 'self_rated_level': 3},
        )
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.self_rated_level, 3)
        self.assertEqual(es.rating_status, 'pending')

    def test_employee_cannot_submit_self_rating_for_someone_else(self):
        self.client.force_login(self.employee_user)
        response = self.client.post(
            reverse('employee_self_rating_update', args=[self.unrelated_employee.id]),
            {'skill_id': self.skill.id, 'self_rated_level': 3},
        )
        self.assertEqual(response.status_code, 403)

    def test_manager_can_approve_own_reports_self_rating(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(reverse('employee_skill_approve', args=[self.employee.id, self.skill.id]))
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.actual_level, 4)
        self.assertEqual(es.rating_status, 'approved')

    def test_non_manager_cannot_approve_a_rating(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        self.client.force_login(self.other_user)
        response = self.client.post(reverse('employee_skill_approve', args=[self.employee.id, self.skill.id]))
        self.assertEqual(response.status_code, 403)

    def test_manager_override_sets_status_overridden(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(
            reverse('employee_skill_update', args=[self.employee.id]),
            {'skill_id': self.skill.id, 'actual_level': 3},
        )
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.actual_level, 3)
        self.assertEqual(es.rating_status, 'overridden')

    def test_my_skills_view_handles_unlinked_account_gracefully(self):
        unlinked_user = User.objects.create_user('unlinked1', password='pass12345')
        self.client.force_login(unlinked_user)
        response = self.client.get(reverse('my_skills'))
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context['employee'])

    def test_approve_blocked_while_development_plan_is_active(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course', status='in_progress',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(reverse('employee_skill_approve', args=[self.employee.id, self.skill.id]))
        self.assertEqual(response.status_code, 409)
        self.assertFalse(response.json()['success'])
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.actual_level, 2)
        self.assertEqual(es.rating_status, 'pending')

    def test_approve_allowed_once_development_plan_is_completed(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course', status='completed',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(reverse('employee_skill_approve', args=[self.employee.id, self.skill.id]))
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.actual_level, 4)
        self.assertEqual(es.rating_status, 'approved')

    def test_manager_entering_matching_level_stays_pending_while_plan_active(self):
        """Even if the manager's own entry happens to match the employee's self-rating,
        that must not silently count as approval while a development plan is still open."""
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course', status='not_started',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(
            reverse('employee_skill_update', args=[self.employee.id]),
            {'skill_id': self.skill.id, 'actual_level': 4},
        )
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.actual_level, 4)
        self.assertEqual(es.rating_status, 'pending')

    def test_cancelled_development_plan_does_not_block_approval(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course', status='cancelled',
        )
        self.client.force_login(self.manager_user)
        response = self.client.post(reverse('employee_skill_approve', args=[self.employee.id, self.skill.id]))
        self.assertEqual(response.status_code, 200)
        es = EmployeeSkill.objects.get(skill_matrix=self.employee, skill=self.skill)
        self.assertEqual(es.rating_status, 'approved')

    def test_has_active_development_plan_false_when_no_plan_exists(self):
        es = EmployeeSkill.objects.create(skill_matrix=self.employee, skill=self.skill, actual_level=2)
        self.assertFalse(es.has_active_development_plan)

    def test_profile_page_shows_blocked_state_for_approve_button(self):
        EmployeeSkill.objects.create(
            skill_matrix=self.employee, skill=self.skill, actual_level=2,
            self_rated_level=4, rating_status='pending',
        )
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course', status='in_progress',
        )
        self.client.force_login(self.manager_user)
        response = self.client.get(reverse('employee_profile', args=[self.employee.id]))
        self.assertContains(response, 'Blocked')
        self.assertContains(response, 'disabled')


class RoleMatrixQuerysetTests(TestCase):
    """get_required_benchmarks()/get_missing_benchmarks() must always return a QuerySet,
    never a plain list, since callers chain .values_list()/.filter() onto the result."""

    def test_get_required_benchmarks_is_chainable_without_a_role(self):
        employee = SkillMatrix.objects.create(name='No Role')
        skill_ids = employee.get_required_benchmarks().values_list('skill_id', flat=True)
        self.assertEqual(list(skill_ids), [])

    def test_get_missing_benchmarks_is_chainable_without_a_role(self):
        employee = SkillMatrix.objects.create(name='No Role')
        self.assertEqual(list(employee.get_missing_benchmarks()), [])


class DevelopmentPlanFormForRoleLessEmployeeTests(TestCase):
    """Regression test: DevelopmentPlanCreateView/UpdateView crashed with AttributeError
    for any employee with no role_matrix, because get_required_benchmarks() used to
    return a plain list and .values_list() was called on it directly."""

    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin3', password='pass12345', is_staff=True)
        self.employee = SkillMatrix.objects.create(name='No Role Employee')
        self.skill = Skill.objects.create(name='Python')

    def test_create_form_loads_for_employee_without_a_role(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse('development_plan_add', args=[self.employee.id]))
        self.assertEqual(response.status_code, 200)

    def test_update_form_loads_for_employee_without_a_role(self):
        self.client.force_login(self.staff_user)
        plan = DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a course',
        )
        response = self.client.get(reverse('development_plan_update', args=[plan.id]))
        self.assertEqual(response.status_code, 200)


class ConfirmDeletePageTests(TestCase):
    """Regression tests: several confirm-delete templates referenced context variables
    Django's DeleteView never actually provides (e.g. {{ employee.name }} when only
    {{ object }} is set), and benchmark_confirm_delete.html used a nonexistent
    `object.designation` field inside a {% url %} tag, which raised NoReverseMatch."""

    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin4', password='pass12345', is_staff=True)
        self.client.force_login(self.staff_user)
        self.role = RoleMatrix.objects.create(title='Engineer')
        self.skill = Skill.objects.create(name='Python')
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=self.role)
        self.benchmark = SkillBenchmark.objects.create(role_matrix=self.role, skill=self.skill, required_level=3)

    def test_employee_delete_confirm_shows_employee_name(self):
        response = self.client.get(reverse('employee_delete', args=[self.employee.id]))
        self.assertContains(response, 'Ada Lovelace')

    def test_designation_delete_confirm_shows_role_title(self):
        response = self.client.get(reverse('designation_delete', args=[self.role.id]))
        self.assertContains(response, 'Engineer')

    def test_benchmark_delete_confirm_renders_without_error(self):
        response = self.client.get(reverse('benchmark_delete', args=[self.benchmark.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Engineer')

    def test_designation_benchmark_delete_confirm_renders_without_error(self):
        response = self.client.get(reverse('designation_benchmark_delete', args=[self.benchmark.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Engineer')

    def test_employee_card_shows_role_title_not_placeholder(self):
        response = self.client.get(reverse('employee_card', args=[self.employee.id]))
        self.assertContains(response, 'Engineer')
        self.assertNotContains(response, 'No designation assigned')

    def test_employee_card_shows_development_plans(self):
        DevelopmentPlan.objects.create(
            skill_matrix=self.employee, skill=self.skill, action='Take a Python course',
            status='in_progress',
        )
        response = self.client.get(reverse('employee_card', args=[self.employee.id]))
        self.assertContains(response, 'Development Plans')
        self.assertContains(response, 'Take a Python course')
        self.assertContains(response, 'In Progress')

    def test_employee_card_hides_development_plans_section_when_none(self):
        response = self.client.get(reverse('employee_card', args=[self.employee.id]))
        self.assertNotContains(response, 'Development Plans')


class SafeJsonTests(TestCase):
    """A skill/employee name containing "</script>" must not be able to close the
    surrounding <script> block early when embedded via {{ ... |safe }}."""

    def test_escapes_script_closing_tag(self):
        rendered = safe_json(['</script><script>alert(1)</script>'])
        self.assertNotIn('</script>', rendered)
        self.assertIn('\\u003c/script\\u003e', rendered)

    def test_still_valid_json_content_for_plain_values(self):
        rendered = safe_json(['Python', 3])
        self.assertIn('"Python"', rendered)
        self.assertIn('3', rendered)


class MalformedInputTests(TestCase):
    """Non-numeric POST values must return a clean error, not an uncaught ValueError."""

    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin5', password='pass12345', is_staff=True)
        self.employee = SkillMatrix.objects.create(name='Ada Lovelace')
        self.skill = Skill.objects.create(name='Python')

    def test_employee_skill_update_rejects_non_numeric_level(self):
        self.client.force_login(self.staff_user)
        response = self.client.post(
            reverse('employee_skill_update', args=[self.employee.id]),
            {'skill_id': self.skill.id, 'actual_level': 'not-a-number'},
        )
        self.assertEqual(response.status_code, 400)

    def test_bulk_skill_update_rejects_non_numeric_level(self):
        self.client.force_login(self.staff_user)
        response = self.client.post(reverse('bulk_skill_update'), {
            'employee_ids': [self.employee.id],
            'skill_id': self.skill.id,
            'actual_level': 'not-a-number',
        })
        self.assertEqual(response.status_code, 302)
        self.assertFalse(EmployeeSkill.objects.filter(skill_matrix=self.employee, skill=self.skill).exists())


class TeamReportDataTests(TestCase):
    def setUp(self):
        self.role = RoleMatrix.objects.create(title='Engineer')
        self.other_role = RoleMatrix.objects.create(title='Manager')
        self.mandatory_skill = Skill.objects.create(name='Python')
        self.other_role_only_skill = Skill.objects.create(name='Leadership')

        SkillBenchmark.objects.create(role_matrix=self.role, skill=self.mandatory_skill, required_level=4, is_mandatory=True)
        SkillBenchmark.objects.create(role_matrix=self.other_role, skill=self.other_role_only_skill, required_level=3, is_mandatory=False)

        self.engineer = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=self.role)
        EmployeeSkill.objects.create(skill_matrix=self.engineer, skill=self.mandatory_skill, actual_level=1)  # gap 3

        self.manager = SkillMatrix.objects.create(name='Grace Hopper', role_matrix=self.other_role)
        EmployeeSkill.objects.create(skill_matrix=self.manager, skill=self.other_role_only_skill, actual_level=3)  # gap 0

    def test_matrix_includes_union_of_skills_across_roles(self):
        data = build_team_report_data(SkillMatrix.objects.all())
        skill_names = {s.name for s in data['skills']}
        self.assertEqual(skill_names, {'Python', 'Leadership'})

    def test_skill_not_benchmarked_for_a_role_is_marked_not_applicable(self):
        data = build_team_report_data(SkillMatrix.objects.all())
        cell = data['matrix'][(self.manager.id, self.mandatory_skill.id)]
        self.assertFalse(cell['applicable'])

    def test_applicable_cell_has_correct_actual_and_required(self):
        data = build_team_report_data(SkillMatrix.objects.all())
        cell = data['matrix'][(self.engineer.id, self.mandatory_skill.id)]
        self.assertEqual(cell, {
            'applicable': True, 'required': 4, 'actual': 1, 'gap': 3,
            'status': 'critical', 'is_mandatory': True,
        })

    def test_summary_totals_match_both_employees(self):
        data = build_team_report_data(SkillMatrix.objects.all())
        self.assertEqual(data['summary']['total_employees'], 2)
        self.assertEqual(data['summary']['total_skills'], 2)
        self.assertEqual(data['summary']['skills_met_pct'], 50)  # 1 of 2 comparisons met

    def test_gap_by_employee_only_includes_employees_with_benchmarks(self):
        unbenched = SkillMatrix.objects.create(name='No Role')
        data = build_team_report_data(SkillMatrix.objects.all())
        names = {e['name'] for e in data['gap_by_employee']}
        self.assertIn('Ada Lovelace', names)
        self.assertNotIn(unbenched.name, names)


class TeamReportExportViewTests(TestCase):
    def setUp(self):
        self.staff_user = User.objects.create_user('hr_admin6', password='pass12345', is_staff=True)
        self.plain_user = User.objects.create_user('plain6', password='pass12345')
        role = RoleMatrix.objects.create(title='Engineer')
        skill = Skill.objects.create(name='Python')
        SkillBenchmark.objects.create(role_matrix=role, skill=skill, required_level=4)
        employee = SkillMatrix.objects.create(name='Ada Lovelace', role_matrix=role)
        EmployeeSkill.objects.create(skill_matrix=employee, skill=skill, actual_level=2)

    def test_anonymous_redirected_from_both_exports(self):
        self.assertEqual(self.client.get(reverse('team_report_excel')).status_code, 302)
        self.assertEqual(self.client.get(reverse('team_report_pdf')).status_code, 302)

    def test_non_staff_forbidden_from_both_exports(self):
        self.client.force_login(self.plain_user)
        self.assertEqual(self.client.get(reverse('team_report_excel')).status_code, 403)
        self.assertEqual(self.client.get(reverse('team_report_pdf')).status_code, 403)

    def test_staff_gets_a_valid_excel_file(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse('team_report_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ['Dashboard', 'Skill Matrix'])
        self.assertEqual(wb['Skill Matrix']['A1'].value, 'Skill')
        self.assertEqual(wb['Skill Matrix']['C1'].value, 'Ada Lovelace')

    def test_staff_gets_a_valid_pdf_file(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(reverse('team_report_pdf'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertGreater(len(response.content), 500)
