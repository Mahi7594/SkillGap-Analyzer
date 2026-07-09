"""Renders the team skill gap report built by reports.build_team_report_data() into
Excel (openpyxl) and PDF (reportlab). Both consume the exact same data structure so the
two formats can never disagree with each other.
"""
from io import BytesIO

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

STATUS_FILL = {
    'met': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
    'warning': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
    'critical': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
}
STATUS_FONT_COLOR = {
    'met': Font(color='059669'),
    'warning': Font(color='D97706'),
    'critical': Font(color='DC2626'),
}
NOT_APPLICABLE_FILL = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_BORDER = Border(*(Side(style='thin', color='E2E8F0'),) * 4)


# --------------------------------------------------------------------------- Excel

def render_team_report_excel(report_data):
    wb = Workbook()
    _write_dashboard_sheet(wb.active, report_data)
    _write_matrix_sheet(wb.create_sheet('Skill Matrix'), report_data)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="team_skill_gap_report.xlsx"'
    return response


def _write_dashboard_sheet(ws, data):
    ws.title = 'Dashboard'
    ws.sheet_view.showGridLines = False

    ws['A1'] = 'Team Skill Gap Report'
    ws['A1'].font = Font(size=18, bold=True, color='1E293B')
    ws.merge_cells('A1:D1')

    summary = data['summary']
    kpi_rows = [
        ('Total Employees', summary['total_employees']),
        ('Skills Tracked', summary['total_skills']),
        ('Avg Gap Score', summary['avg_gap_score']),
        ('Skills Met %', f"{summary['skills_met_pct']}%"),
    ]
    row = 3
    for label, value in kpi_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color='64748B')
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = Font(size=14, bold=True, color='4F46E5')
        row += 1

    row += 1
    row = _write_chart_block(ws, row, 'Average Gap by Role', ['Role', 'Avg Gap', 'Employees'],
                              [[r['role'], r['avg_gap'], r['employee_count']] for r in data['gap_by_role']])
    row = _write_chart_block(ws, row, 'Top Skills Needing Training', ['Skill', 'Avg Gap'],
                              [[s['name'], s['avg_gap']] for s in data['gap_by_skill']])
    _write_chart_block(ws, row, 'Gap by Employee', ['Employee', 'Avg Gap'],
                        [[e['name'], e['avg_gap']] for e in data['gap_by_employee']])

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14


def _write_chart_block(ws, start_row, title, headers, rows):
    """Writes a small labeled table starting at start_row, then a bar chart of its
    first numeric column next to it. Returns the row number to continue writing at."""
    ws.cell(row=start_row, column=1, value=title).font = Font(size=13, bold=True, color='1E293B')
    header_row = start_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    if not rows:
        ws.cell(row=header_row + 1, column=1, value='No data').font = Font(italic=True, color='94A3B8')
        return header_row + 3

    for i, values in enumerate(rows):
        for col, value in enumerate(values, start=1):
            ws.cell(row=header_row + 1 + i, column=col, value=value)

    last_data_row = header_row + len(rows)

    chart = BarChart()
    chart.type = 'bar'
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.height = 6
    chart.width = 16
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=header_row, max_row=last_data_row)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    ws.add_chart(chart, f'E{start_row}')

    return last_data_row + 3


def _write_matrix_sheet(ws, data):
    ws.sheet_view.showGridLines = False
    employees = data['employees']
    skills = data['skills']
    matrix = data['matrix']

    ws.cell(row=1, column=1, value='Skill').font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value='Mandatory').font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL

    for col, emp in enumerate(employees, start=3):
        name_cell = ws.cell(row=1, column=col, value=emp.name)
        name_cell.font = HEADER_FONT
        name_cell.fill = HEADER_FILL
        name_cell.alignment = Alignment(horizontal='center')
        role_cell = ws.cell(row=2, column=col, value=emp.role_matrix.title if emp.role_matrix else '-')
        role_cell.font = Font(italic=True, color='64748B')
        role_cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12

    for row_offset, skill in enumerate(skills):
        row = row_offset + 3
        skill_cell = ws.cell(row=row, column=1, value=skill.name)
        skill_cell.font = Font(bold=True)

        any_mandatory = any(
            matrix.get((emp.id, skill.id), {}).get('applicable') and matrix[(emp.id, skill.id)]['is_mandatory']
            for emp in employees
        )
        ws.cell(row=row, column=2, value='Yes' if any_mandatory else 'Optional')

        for col, emp in enumerate(employees, start=3):
            cell_data = matrix.get((emp.id, skill.id), {'applicable': False})
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            if not cell_data['applicable']:
                cell.value = '—'
                cell.fill = NOT_APPLICABLE_FILL
                cell.font = Font(color='CBD5E1')
            else:
                cell.value = f"{cell_data['actual']}/{cell_data['required']}"
                cell.fill = STATUS_FILL[cell_data['status']]
                cell.font = STATUS_FONT_COLOR[cell_data['status']]

    ws.freeze_panes = 'C3'


# --------------------------------------------------------------------------- PDF

EMPLOYEES_PER_PDF_CHUNK = 8


def render_team_report_pdf(report_data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], textColor=colors.HexColor('#1E293B'))
    heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], textColor=colors.HexColor('#1E293B'),
                                    spaceBefore=14, spaceAfter=6)

    story = [Paragraph('Team Skill Gap Report', title_style), Spacer(1, 0.4 * cm)]
    story.extend(_pdf_summary_table(report_data['summary'], styles))
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph('Average Gap by Role', heading_style))
    story.append(_pdf_bar_chart(
        [r['role'] for r in report_data['gap_by_role']],
        [r['avg_gap'] for r in report_data['gap_by_role']],
    ))

    story.append(Paragraph('Top Skills Needing Training', heading_style))
    story.append(_pdf_bar_chart(
        [s['name'] for s in report_data['gap_by_skill']],
        [s['avg_gap'] for s in report_data['gap_by_skill']],
    ))

    story.append(Paragraph('Gap by Employee', heading_style))
    story.append(_pdf_bar_chart(
        [e['name'] for e in report_data['gap_by_employee']],
        [e['avg_gap'] for e in report_data['gap_by_employee']],
    ))

    story.append(Paragraph('Skill Matrix (Actual / Required)', heading_style))
    story.extend(_pdf_matrix_tables(report_data))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="team_skill_gap_report.pdf"'
    return response


def _pdf_summary_table(summary, styles):
    data = [
        ['Total Employees', 'Skills Tracked', 'Avg Gap Score', 'Skills Met %'],
        [summary['total_employees'], summary['total_skills'], summary['avg_gap_score'], f"{summary['skills_met_pct']}%"],
    ]
    table = Table(data, colWidths=[6 * cm] * 4)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 16),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    return [table]


def _pdf_bar_chart(labels, values):
    if not labels:
        return Spacer(1, 0.2 * cm)

    drawing = Drawing(24 * cm, 6 * cm)
    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 30
    chart.height = 5 * cm
    chart.width = 22 * cm
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.dy = -10
    chart.categoryAxis.labels.fontSize = 7
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor('#4F46E5')
    drawing.add(chart)
    return drawing


def _pdf_matrix_tables(report_data):
    employees = report_data['employees']
    skills = report_data['skills']
    matrix = report_data['matrix']
    styles = getSampleStyleSheet()
    note_style = ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#64748B'))

    flowables = []
    for chunk_start in range(0, len(employees), EMPLOYEES_PER_PDF_CHUNK):
        chunk = employees[chunk_start:chunk_start + EMPLOYEES_PER_PDF_CHUNK]

        header_row = ['Skill'] + [emp.name for emp in chunk]
        role_row = [''] + [emp.role_matrix.title if emp.role_matrix else '-' for emp in chunk]
        rows = [header_row, role_row]

        for skill in skills:
            row = [skill.name]
            for emp in chunk:
                cell_data = matrix.get((emp.id, skill.id), {'applicable': False})
                row.append(f"{cell_data['actual']}/{cell_data['required']}" if cell_data['applicable'] else '—')
            rows.append(row)

        col_widths = [4.5 * cm] + [((24 * cm) - 4.5 * cm) / len(chunk)] * len(chunk)
        table = Table(rows, colWidths=col_widths, repeatRows=2)
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 7),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#64748B')),
            ('FONTNAME', (0, 2), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTSIZE', (0, 2), (-1, -1), 8),
        ]
        for row_idx, skill in enumerate(skills, start=2):
            for col_idx, emp in enumerate(chunk, start=1):
                cell_data = matrix.get((emp.id, skill.id), {'applicable': False})
                if not cell_data['applicable']:
                    style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#CBD5E1')))
                elif cell_data['status'] == 'met':
                    style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#D1FAE5')))
                elif cell_data['status'] == 'warning':
                    style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FEF3C7')))
                else:
                    style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FEE2E2')))
        table.setStyle(TableStyle(style_commands))

        flowables.append(table)
        remaining = len(employees) - (chunk_start + len(chunk))
        if remaining > 0:
            flowables.append(Paragraph(f'({remaining} more employee{"s" if remaining != 1 else ""} below)', note_style))
        flowables.append(Spacer(1, 0.4 * cm))

    return flowables
